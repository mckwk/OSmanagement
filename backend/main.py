from fastapi import FastAPI, HTTPException, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import sqlite3
import os
import shutil
import logging

app = FastAPI()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Allow CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Connect to SQLite database
conn = sqlite3.connect('my_database.db', check_same_thread=False)
cursor = conn.cursor()

# Create tables if they don't exist
cursor.execute('''CREATE TABLE IF NOT EXISTS students (
                  id INTEGER PRIMARY KEY,
                  name TEXT,
                  surname TEXT,
                  faculty TEXT,
                  student_id TEXT,
                  phone TEXT,
                  email TEXT,
                  joining_date TEXT,
                  position TEXT
                  )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS activities (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  coordinator_name TEXT,
                  activity_name TEXT,
                  participation_form TEXT,
                  organizer_name TEXT,
                  category TEXT,
                  num_participants TEXT,
                  event_date_from TEXT,
                  event_date_to TEXT,
                  event_scope TEXT,
                  event_type TEXT,
                  description TEXT
                  )''')
conn.commit()

@app.post("/add_student")
async def add_student(
    name: str = Form(...),
    surname: str = Form(...),
    faculty: str = Form(...),
    student_id: str = Form(...),
    phone: str = Form(...),
    email: str = Form(...),
    joining_date: str = Form(...),
    position: str = Form(...)
):
    try:
        logger.info(f"Adding student: {name} {surname}")
        cursor.execute('''INSERT INTO students (name, surname, faculty, student_id, phone, email, joining_date, position)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                       (name, surname, faculty, student_id, phone, email, joining_date, position))
        conn.commit()
        logger.info("Student added successfully")
        return {
            "message": "Student added successfully",
            "name": name,
            "surname": surname,
            "faculty": faculty,
            "student_id": student_id,
            "phone": phone,
            "email": email,
            "joining_date": joining_date,
            "position": position
        }
    except Exception as e:
        logger.error(f"Error adding student: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred while adding the student: {str(e)}")

@app.post("/add_activity")
async def add_activity(
    coordinator_name: str = Form(...),
    activity_name: str = Form(...),
    participation_form: str = Form(...),
    organizer_name: str = Form(...),
    category: str = Form(...),
    num_participants: str = Form(...),
    event_date_from: str = Form(...),
    event_date_to: str = Form(...),
    event_scope: str = Form(...),
    event_type: str = Form(...),
    description: str = Form(...)
):
    try:
        logger.info(f"Adding activity: {activity_name}")
        cursor.execute('''INSERT INTO activities (coordinator_name, activity_name, participation_form, organizer_name, category, num_participants, event_date_from, event_date_to, event_scope, event_type, description)
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (coordinator_name, activity_name, participation_form, organizer_name, category, num_participants, event_date_from, event_date_to, event_scope, event_type, description))
        conn.commit()
        logger.info("Activity added successfully")
        return {
            "message": "Activity added successfully",
            "coordinator_name": coordinator_name,
            "activity_name": activity_name,
            "participation_form": participation_form,
            "organizer_name": organizer_name,
            "category": category,
            "num_participants": num_participants,
            "event_date_from": event_date_from,
            "event_date_to": event_date_to,
            "event_scope": event_scope,
            "event_type": event_type,
            "description": description
        }
    except Exception as e:
        logger.error(f"Error adding activity: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred while adding the activity: {str(e)}")

@app.get("/activities")
async def get_activities():
    try:
        cursor.execute('''SELECT * FROM activities''')
        activities = cursor.fetchall()
        return JSONResponse(content={"activities": activities})
    except Exception as e:
        logger.error(f"Error fetching activities: {e}")
        raise HTTPException(status_code=500, detail="An error occurred while fetching activities")

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    try:
        logger.info(f"Generating report and downloading file: {file_name}")
        
        # Ensure the destination file is a copy of the template
        shutil.copy("formatka.xlsx", file_name)
        
        # Load existing Excel file
        wb = load_workbook(file_name)
        ws = wb.active

        # Retrieve data from SQLite database
        cursor.execute('''SELECT name || " " || surname as FullName, faculty, student_id FROM students''')
        rows = cursor.fetchall()

        # Set starting position
        start_row = 4
        start_col = 2

        # Write data from database to existing Excel file
        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

        # Save changes to the file
        wb.save(file_name)

        file_path = file_name  # Path to your Excel file, adjust accordingly
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File '{file_name}' not found at path: {file_path}")
        elif not os.path.isfile(file_path):
            raise HTTPException(status_code=404, detail=f"Path '{file_path}' is not a file")

        logger.info(f"File {file_name} generated and ready for download")
        return FileResponse(file_path, filename=file_name)
    except Exception as e:
        logger.error(f"Error generating membership report: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred while generating the membership report: {str(e)}")

@app.get("/downloadActivities/{file_name}")
async def download_file(file_name: str):
    try:
        logger.info(f"Generating report and downloading file: {file_name}")
        
        # Ensure the destination file is a copy of the template
        shutil.copy("projekty_i_aktywnosci.xlsx", file_name)
        
        # Load existing Excel file
        wb = load_workbook(file_name)
        ws = wb.active

        # Retrieve data from SQLite database
        cursor.execute('''SELECT coordinator_name, activity_name, participation_form, organizer_name, '' as organizer_choice, category, num_participants, event_date_from, event_date_to, event_scope, event_type, '' as A, '' as B, '' as C, '' as D, '' as E, '' as F, '' as G, '' as H, '' as I, '' as J , description FROM activities''')
        rows = cursor.fetchall()

        # Set starting position
        start_row = 5
        start_col = 2

        # Write data from database to existing Excel file
        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

        # Save changes to the file
        wb.save(file_name)

        file_path = file_name  # Path to your Excel file, adjust accordingly
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File '{file_name}' not found at path: {file_path}")
        elif not os.path.isfile(file_path):
            raise HTTPException(status_code=404, detail=f"Path '{file_path}' is not a file")

        logger.info(f"File {file_name} generated and ready for download")
        return FileResponse(file_path, filename=file_name)
    except Exception as e:
        logger.error(f"Error generating activities report: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred while generating the activities report: {str(e)}")


@app.get("/students")
async def get_students():
    try:
        cursor.execute('''SELECT * FROM students''')
        students = cursor.fetchall()
        return JSONResponse(content={"students": students})
    except Exception as e:
        logger.error(f"Error fetching students: {e}")
        raise HTTPException(status_code=500, detail="An error occurred while fetching students")

@app.get("/students/{student_id}")
async def get_student(student_id: str):
    try:
        cursor.execute('''SELECT * FROM students WHERE student_id = ?''', (student_id,))
        student = cursor.fetchone()
        if student is None:
            raise HTTPException(status_code=404, detail="Student not found")
        return JSONResponse(content={"student": student})
    except Exception as e:
        logger.error(f"Error fetching student {student_id}: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred while fetching the student: {str(e)}")

# Make sure to close the database connection when the application is shut down
@app.on_event("shutdown")
def shutdown_event():
    conn.close()
    logger.info("Database connection closed")